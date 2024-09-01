import os
import nptdms
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import filedialog, Tk
import ctypes



root = Tk()
root.withdraw()
root.attributes('-topmost', True)
root.wm_attributes('-topmost', 1)  
root.lift()
folder_selected = filedialog.askdirectory(title='Directory')
if folder_selected:
    
    print(f'Folder Select:{folder_selected}')
    full_path = os.path.abspath(folder_selected)
    print(full_path)
    file_list = []
    os.system(f'explorer "{full_path}"')
    for root_dir, _, filenames in os.walk(folder_selected):
        for filename in filenames:
            if filename.endswith('.tdms'):
                file_list.append(os.path.join(root_dir, filename))

    num_len=len(file_list)
    print(num_len)              
    for a in range(num_len):
        n=file_list[a]
        print(n)
        tdms_file = nptdms.TdmsFile.read(n)
        xlsx_file = Workbook()
        for group in tdms_file.groups():
            sheet = xlsx_file.create_sheet(title=group.name)
            for i, channel in enumerate(group.channels()):
                column = get_column_letter(i+1)
                sheet[column + '1'] = channel.name
                for j, value in enumerate(channel[:]):
                    sheet[column + str(j+2)] = value
        xlsx_file.save(n+".xlsx")
    
    ctypes.windll.user32.MessageBoxW(0, "Converted data", "Completed Process", 0x40 | 0x1)
else: full_path=None




